import pdfplumber
import pandas as pd
import re
from collections import defaultdict

def extract_transactions_from_pdf(pdf_path, output_excel_path):
    """
    Extract transaction details from bank statement PDF
    """
    
    transactions = []
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Processing {len(pdf.pages)} page(s)...\n")
        
        for page_num, page in enumerate(pdf.pages, 1):
            print(f"--- Page {page_num} ---")
            
            # Method 1: Use implicit table detection (best for tables with clear structure)
            tables = page.extract_tables({
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
                "edge_min_length": 3,
                "min_words_vertical": 3,
                "min_words_horizontal": 1,
                "intersection_tolerance": 15,
            })
            
            print(f"Found {len(tables)} table(s)")
            
            for table_idx, table in enumerate(tables):
                if not table or len(table) < 1:
                    continue
                
                print(f"\nTable {table_idx + 1}: {len(table)} rows, {len(table[0]) if table else 0} columns")
                
                # Check if this is a transaction table (look for DATE column)
                is_transaction_table = False
                header_idx = -1
                
                for i, row in enumerate(table[:3]):  # Check first 3 rows
                    if row:
                        row_str = ' '.join([str(cell).upper() if cell else '' for cell in row])
                        if 'DATE' in row_str and 'PARTICULARS' in row_str:
                            is_transaction_table = True
                            header_idx = i
                            print(f"  Transaction table header at row {i}")
                            break
                
                if not is_transaction_table:
                    print(f"  Skipping - not a transaction table")
                    continue
                
                # Extract transactions
                rows_extracted = 0
                
                for row_idx in range(header_idx + 1, len(table)):
                    row = table[row_idx]
                    
                    if not row or len(row) < 3:
                        continue
                    
                    # Check if first column contains a date
                    first_cell = str(row[0]) if row[0] else ""
                    date_match = re.search(r'\d{2}-\d{2}-\d{4}', first_cell)
                    
                    if not date_match:
                        continue
                    
                    date = date_match.group()
                    
                    # Extract other columns
                    mode = ""
                    particulars = ""
                    deposits = ""
                    withdrawals = ""
                    balance = ""
                    
                    if len(row) >= 2:
                        mode = str(row[1]) if row[1] and str(row[1]) != 'None' else ""
                        mode = mode.replace('\n', ' ').strip()
                    
                    if len(row) >= 3:
                        particulars = str(row[2]) if row[2] and str(row[2]) != 'None' else ""
                        particulars = particulars.replace('\n', ' ').strip()
                    
                    if len(row) >= 4:
                        deposits = str(row[3]) if row[3] and str(row[3]) != 'None' else ""
                        deposits = deposits.replace('\n', ' ').strip()
                    
                    if len(row) >= 5:
                        withdrawals = str(row[4]) if row[4] and str(row[4]) != 'None' else ""
                        withdrawals = withdrawals.replace('\n', ' ').strip()
                    
                    if len(row) >= 6:
                        balance = str(row[5]) if row[5] and str(row[5]) != 'None' else ""
                        balance = balance.replace('\n', ' ').strip()
                    
                    transaction = {
                        'DATE': date,
                        'MODE': mode,
                        'PARTICULARS': particulars,
                        'DEPOSITS': deposits,
                        'WITHDRAWALS': withdrawals,
                        'BALANCE': balance
                    }
                    
                    transactions.append(transaction)
                    rows_extracted += 1
                    
                    # Debug: show first 3
                    if rows_extracted <= 3:
                        print(f"    {date} | {particulars[:50]}...")
                
                print(f"  Extracted {rows_extracted} transactions")
    
    # If no transactions found, try debugging
    if len(transactions) == 0:
        print("\n" + "="*80)
        print("NO TRANSACTIONS FOUND - DEBUGGING")
        print("="*80)
        
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            
            print("\n1. Checking table structure on page 1...")
            tables = page.extract_tables()
            
            print(f"\nFound {len(tables)} tables using default settings")
            
            for i, table in enumerate(tables[:3]):  # Show first 3 tables
                if table:
                    print(f"\n--- Table {i+1} ---")
                    print(f"Dimensions: {len(table)} rows x {len(table[0]) if table[0] else 0} columns")
                    print("\nFirst 5 rows:")
                    for j, row in enumerate(table[:5]):
                        print(f"Row {j}: {row}")
            
            print("\n" + "="*80)
            print("2. Trying different extraction settings...")
            print("="*80)
            
            # Try with text-based strategy
            print("\nMethod A: Text-based extraction")
            tables_text = page.extract_tables({
                "vertical_strategy": "text",
                "horizontal_strategy": "text",
            })
            print(f"Found {len(tables_text)} tables")
            if tables_text and tables_text[0]:
                print(f"First table: {len(tables_text[0])} rows")
                for j, row in enumerate(tables_text[0][:3]):
                    print(f"  Row {j}: {row}")
            
            # Try with explicit lines
            print("\nMethod B: Explicit lines")
            tables_explicit = page.extract_tables({
                "vertical_strategy": "lines_strict",
                "horizontal_strategy": "lines_strict",
            })
            print(f"Found {len(tables_explicit)} tables")
            if tables_explicit and tables_explicit[0]:
                print(f"First table: {len(tables_explicit[0])} rows")
                for j, row in enumerate(tables_explicit[0][:3]):
                    print(f"  Row {j}: {row}")
            
            # Try extracting words to understand structure
            print("\n" + "="*80)
            print("3. Analyzing text structure...")
            print("="*80)
            
            words = page.extract_words()
            print(f"\nFound {len(words)} words on page 1")
            
            # Find words containing "DATE"
            date_words = [w for w in words if 'DATE' in w['text'].upper()]
            if date_words:
                print(f"\nFound 'DATE' at position: x={date_words[0]['x0']:.1f}, y={date_words[0]['top']:.1f}")
            
            # Find first date pattern
            date_pattern_words = [w for w in words if re.match(r'\d{2}-\d{2}-\d{4}', w['text'])]
            if date_pattern_words:
                print(f"Found first date '{date_pattern_words[0]['text']}' at: x={date_pattern_words[0]['x0']:.1f}, y={date_pattern_words[0]['top']:.1f}")
                
                # Show nearby words
                first_date_y = date_pattern_words[0]['top']
                nearby_words = [w for w in words if abs(w['top'] - first_date_y) < 20]
                nearby_words.sort(key=lambda w: w['x0'])
                
                print("\nWords near first date (sorted left to right):")
                for w in nearby_words[:10]:
                    print(f"  x={w['x0']:6.1f}: '{w['text']}'")
            
            # Show raw text
            print("\n" + "="*80)
            print("4. Raw text (first 2000 chars)...")
            print("="*80)
            text = page.extract_text()
            print(text[:2000])
        
        return None
    
    # Create DataFrame
    df = pd.DataFrame(transactions)
    
    # Clean data
    df = df.replace('None', '')
    df = df.drop_duplicates()
    df = df.reset_index(drop=True)
    
    # Save to Excel
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Transactions')
        
        worksheet = writer.sheets['Transactions']
        
        # Column widths
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 70
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        
        # Formatting
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            row[0].alignment = Alignment(horizontal='left', vertical='top')
            row[1].alignment = Alignment(horizontal='left', vertical='top')
            row[2].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            row[3].alignment = Alignment(horizontal='right', vertical='top')
            row[4].alignment = Alignment(horizontal='right', vertical='top')
            row[5].alignment = Alignment(horizontal='right', vertical='top')
            
            for cell in row:
                cell.border = thin_border
    
    print(f"\n✅ Successfully extracted {len(df)} transactions")
    print(f"✅ Excel file saved: {output_excel_path}")
    
    return df

if __name__ == "__main__":
    pdf_file = "transaction.pdf"
    excel_file = "bank_transactions_final1121.xlsx"
    
    try:
        df = extract_transactions_from_pdf(pdf_file, excel_file)
        
        if df is not None and len(df) > 0:
            print("\n" + "="*100)
            print("EXTRACTED TRANSACTIONS")
            print("="*100)
            
            if len(df) <= 15:
                print(df.to_string(index=False, max_colwidth=60))
            else:
                print("\nFirst 10:")
                print(df.head(10).to_string(index=False, max_colwidth=60))
                print(f"\n... {len(df)-15} more transactions ...")
                print("\nLast 5:")
                print(df.tail(5).to_string(index=False, max_colwidth=60))
            
            print("\n" + "="*100)
            print(f"✅ Total: {len(df)} transactions")
            print(f"✅ Period: {df['DATE'].iloc[0]} to {df['DATE'].iloc[-1]}")
        
    except FileNotFoundError:
        print(f"❌ PDF file not found: '{pdf_file}'")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()